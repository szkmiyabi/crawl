import requests
import lxml.html
import re
import datetime
import time
import html
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

class CSSReportBots:
    def __init__(self, urls_filename):
        self.urls_filename = urls_filename
        self.shortWait = 1
        self.midWait = 3
        self.longWait = 10
        self.max_row_cnt = 0
    
    def load_url_datas(self):
        datas = []
        with open(self.urls_filename, "r") as f:
            for r in f:
                line = r.strip()
                cols = line.split("\t")
                datas.append({"pid": cols[0], "url": cols[1]})
        return datas

    def fetch_css_datas(self, cr_pid, cr_url):
        print(cr_pid, "を処理しています")
        time.sleep(self.shortWait)
        datas = []
        try:
            res = requests.get(cr_url, verify=False)
            doc = res.text
            dom = lxml.html.fromstring(res.content)
            dom.make_links_absolute(res.url)
            link_tags = []
            style_tags = []
            style_atts = ""
            for link in dom.cssselect("link"):
                if link.get("rel") == "stylesheet":
                    link_tags.append(link.get("href"))
            for style in dom.cssselect("style"):
                style_str = html.unescape(lxml.html.tostring(style).decode())
                style_tags.append(style_str)
            for style_attr in re.findall(r'style=".*?"', doc, re.DOTALL):
                style_atts += style_attr + ","
            datas.append({
                "pid": cr_pid,
                "url": cr_url,
                "link_tags": link_tags,
                "style_tags": style_tags,
                "style_atts": style_atts
            })
        except requests.exceptions.SSLError:
            print(cr_pid, "はエラーのためスキップします")
        return datas
    
    def is_exist_solid_size(self, cr_text):
        res = self.text_to_linear(cr_text)
        if re.search(r'(font-size: *?[0-9]+px;|font-size: *?[0-9]+pt;)', res, re.DOTALL):
            return True
        else:
            return False
    
    def is_exist_font_size(self, cr_text):
        res = self.text_to_linear(cr_text)
        if re.search(r'(font-size: *.+?;)', res, re.DOTALL):
            return True
        else:
            return False
    
    def fetch_row_data(self, cr_datas):
        time.sleep(self.shortWait)
        datas = {}
        for r in cr_datas:
            pid = r["pid"]
            url = r["url"]
            link_tags = r["link_tags"]
            style_tags = r["style_tags"]
            style_atts = r["style_atts"]
            link_tags_res = []
            style_tags_res = []
            style_atts_res = ""
            datas.update({"pid": pid, "url": url})
            for link_tags_r in link_tags:
                tmp = self.cssfile_request(link_tags_r)
                if self.is_exist_solid_size(tmp) is True:
                    link_tags_res.append("NG")
                elif self.is_exist_font_size(tmp) is True:
                    link_tags_res.append("OK")
                else:
                    link_tags_res.append("NA")
            for style_tags_r in style_tags:
                if self.is_exist_solid_size(style_tags_r) is True:
                    style_tags_res.append("NG")
                elif self.is_exist_font_size(style_tags_r) is True:
                    style_tags_res.append("OK")
                else:
                    style_tags_res.append("NA")
            if self.is_exist_solid_size(style_atts) is True:
                style_atts_res = "NG"
            elif self.is_exist_font_size(style_atts) is True:
                style_atts_res = "OK"
            else:
                style_atts_res = "NA"
            datas.update({
                "link_tags": link_tags,
                "link_tags_res": link_tags_res,
                "style_tags": style_tags,
                "style_tags_res": style_tags_res,
                "style_atts": style_atts,
                "style_atts_res": style_atts_res
            })
        return datas

    def cssfile_request(self, cr_url):
        docall = ""
        try:
            res = requests.get(cr_url, verify=False)
            doc = res.text
            docall += doc
            if self.is_exist_import_css(doc) is True:
                for r in self.get_import_css_list(cr_url, doc):
                    try:
                        time.sleep(self.shortWait)
                        rres = requests.get(r, verify=False)
                        rdoc = rres.text
                        docall += rdoc
                    except requests.exceptions.SSLError:
                        print("@import CSSの取得エラーです")
        except requests.exceptions.SSLError:
            print("外部CSS取得エラーです")
        return docall
    
    def text_to_linear(self, cr_text):
        return re.sub(r'(\r\n|\r|\n|\s{2,})', "", cr_text)

    def is_exist_import_css(self, cr_text):
        if re.search(r'@import url.+?;', cr_text, re.DOTALL):
            return True
        else:
            return False
    
    def get_import_css_list(self, cr_css_url, cr_text):
        datas = []
        domain = self.get_domain(cr_css_url)
        for partial_code in re.findall(r'@import url\("*.+?"*\);', cr_text, re.DOTALL):
            mt = re.search(r'(@import url\("*)(.+?)("*\);)', partial_code, re.DOTALL).group(2)
            print(mt)
            datas.append(domain + mt)
        return datas

    def get_domain(self, cr_url):
        return re.search(r'(http.*://.+?/)(.+)', cr_url, re.DOTALL).group(1)
    
    def save_as_xlsx(self, datas):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "ページID"
        ws.cell(row=1, column=2).value = "URL"
        ws.cell(row=1, column=3).value = "外部CSS"
        ws.cell(row=1, column=4).value = "固定size"
        ws.cell(row=1, column=5).value = "style属性"
        ws.cell(row=1, column=6).value = "固定size"
        ws.cell(row=1, column=7).value = "style属性"
        ws.cell(row=1, column=8).value = "固定size"

        row = 2
        for r in datas:
            pid = r["pid"]
            url = r["url"]
            link_tags = r["link_tags"]
            link_tags_res = r["link_tags_res"]
            style_tags = r["style_tags"]
            style_tags_res = r["style_tags_res"]
            style_atts = r["style_atts"]
            style_atts_res = r["style_atts_res"]
            col = 1
            row_cnt = 0
            ws.cell(row=row, column=col).value = pid
            col += 1
            ws.cell(row=row, column=col).value = url
            col += 1
            for link in link_tags:
                ws.cell(row=row, column=col).value = link
                row += 1
                row_cnt += 1
                self.regist_max(row_cnt)
            col += 1
            row_cnt_link_res = 0
            for link_res in link_tags_res:
                if row_cnt_link_res == 0:
                    row -= row_cnt
                    row_cnt =0
                else:
                    pass
                ws.cell(row=row, column=col).value = link_res
                row += 1
                row_cnt += 1
                row_cnt_link_res += 1
                self.regist_max(row_cnt)
            col += 1
            row_cnt_style = 0
            for style in style_tags:
                if row_cnt_style == 0:
                    row -= row_cnt
                    row_cnt = 0
                else:
                    pass
                ws.cell(row=row, column=col).value = style
                row += 1
                row_cnt += 1
                row_cnt_style += 1
                self.regist_max(row_cnt)
            if len(style_tags) == 0:
                row -= row_cnt
                row_cnt = 0
                ws.cell(row=row, column=col).value = "存在しません"
            else:
                pass
            col += 1
            row_cnt_style_res = 0
            for style_res in style_tags_res:
                if row_cnt_style_res == 0:
                    row -= row_cnt
                    row_cnt = 0
                else:
                    pass
                ws.cell(row=row, column=col).value = style_res
                row += 1
                row_cnt += 1
                row_cnt_style_res += 1
                self.regist_max(row_cnt)
            col += 1
            row -= row_cnt
            ws.cell(row=row, column=col).value = style_atts
            if len(style_atts) == 0:
                ws.cell(row=row, column=col).value = "存在しません"
            else:
                pass
            col += 1
            ws.cell(row=row, column=col).value = style_atts_res
            row += self.max_row_cnt
            self.reset_max()
        wb.save(self.fetch_filename_from_datetime(".xlsx"))

    def regist_max(self, val):
        if val > self.max_row_cnt:
            self.max_row_cnt = val
        else:
            pass
    
    def reset_max(self):
        self.max_row_cnt = 0

    def fetch_filename_from_datetime(self, ext_str):
        datetime_fmt = datetime.datetime.today()
        return datetime_fmt.strftime("%Y-%m-%d_%H-%M-%S") + ext_str
    
    def exec(self):
        parent_datas = []
        datas = self.load_url_datas()
        for r in datas:
            pid = r["pid"]
            url = r["url"]
            in_datas = self.fetch_css_datas(pid, url)
            res_data = self.fetch_row_data(in_datas)
            print(res_data)
            parent_datas.append(res_data)
        self.save_as_xlsx(parent_datas)

app = CSSReportBots("urls.txt")
app.exec()
        
