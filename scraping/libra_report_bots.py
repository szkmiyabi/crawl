from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import sys
import datetime
import time
import yaml
import lxml.html
import html
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

class LibraRepBots:
    def __init__(self, projectID):
        with open("user.yaml") as f:
            userdata = yaml.load(f)
            self.uid = userdata["uid"]
            self.passwd = userdata["pswd"]
            self.systemWait = userdata["systemWait"]
            self.longWait = userdata["longWait"]
            self.midWait = userdata["midWait"]
            self.shortWait = userdata["shortWait"]
            self.hrefFlag = userdata["hrefFlag"]
        self.app_url = "https://accessibility.jp/libra/"
        self.index_url = "https://jis.infocreate.co.jp/"
        self.rep_index_url_base = "http://jis.infocreate.co.jp/diagnose/indexv2/report/projID/"
        self.rep_detail_url_base = "http://jis.infocreate.co.jp/diagnose/indexv2/report2/projID/"
        self.sv_mainpage_url_base = "http://jis.infocreate.co.jp/diagnose/indexv2/index/projID/"
        self.guideline_filename = "guideline_datas.txt"
        self.projectID = projectID
        self.wd = webdriver.PhantomJS()
        self.wd.implicitly_wait(self.systemWait)
        self.wd.set_window_size(1280, 900)
        self.wd.get(self.app_url)
    
    def getWd(self):
        return self.wd
    
    def login(self):
        self.wd.find_element_by_id("uid").send_keys(self.uid)
        self.wd.find_element_by_id("pswd").send_keys(self.passwd)
        self.wd.find_element_by_id("btn").click()

    def logout(self):
        self.wd.get(self.index_url)
        btnWrap = wd.find_element_by_id("btn")
        btnBase = btnWrap.find_element_by_tag_name("ul")
        btnBaseInner = btnBase.find_element_by_class_name("btn2")
        btnA = btnBaseInner.find_element_by_tag_name("a")
        btnA.click()
    
    def shutdown(self):
        self.wd.quit()
    
    def save_sc(self, filename):
        self.wd.save_screenshot(filename)

    def browse_rep(self):
        self.wd.get(self.rep_index_url_base + self.projectID)

    def fetch_filename_from_datetime(self, ext_str):
        datetime_fmt = datetime.datetime.today()
        return datetime_fmt.strftime("%Y-%m-%d_%H-%M-%S") + ext_str
    
    def fetch_report_detail_path(self, pageID, guidelineID):
        return self.rep_detail_url_base + self.projectID + "/controlID/" + pageID + "/guideline/" + guidelineID + "/"

    def fetch_sv_mainpage_path(self, pageID):
        return self.sv_mainpage_url_base + self.projectID + "/controlID/" + '"' + pageID + '"'

    def get_dom(self):
        html_str = self.wd.page_source
        return lxml.html.fromstring(html_str)
    
    def get_page_list_data(self):
        datas = []
        dom = self.get_dom()
        tbl = dom.xpath("/html/body/div[2]/div[1]/table")[0]
        for row in tbl.cssselect("tr td:first-child"):
            td_val = row.text
            datas.append(td_val)
        return datas
    
    def get_detail_table_data(self, pageID, guideline):
        datas = []
        dom = self.get_dom()
        tbl = dom.xpath("/html/body/div[2]/div[2]/table")[0]
        cnt = 0
        for row in tbl.cssselect("tr"):
            cnt += 1
            if cnt < 2:
                pass
            else:
                tr = lxml.html.tostring(row).decode()
                trdom = lxml.html.fromstring(tr)
                row_datas = []
                row_datas.append(pageID)
                row_datas.append(guideline)
                if self.hrefFlag == "yes":
                    row_datas.append(self.fetch_sv_mainpage_path(pageID))
                else:
                    pass
                for irow in trdom.cssselect("td"):
                    td_val = irow.text_content()
                    if td_val is None:
                        if self.is_including_tag(irow) is True:
                            row_datas.append(self.get_regx_text(irow))
                        else:
                            row_datas.append("")
                    else:
                        row_datas.append(td_val.strip())
                datas.append(row_datas)
        return datas

    def is_including_tag(self, elm):
        tag_str = html.unescape(lxml.html.tostring(elm).decode())
        if re.search(r'<td.*?>(.+)</td>', tag_str.strip(), re.DOTALL):
            return True
        else:
            return False

    def get_regx_text(self, elm):
        tag_str = html.unescape(lxml.html.tostring(elm).decode())
        tag_str = re.sub(r'(\r|\n|\r\n|\t|\s{2,})', "", tag_str)
        return re.search(r'<td.*?>(.+)</td>', tag_str.strip(), re.DOTALL).group(1)

    def open_text_data(self, filename):
        line = []
        with open(filename) as f:
            line = [s.strip() for s in f.readlines()]
        return line
    
    def save_text(self, text_data):
        with open(self.fetch_filename_from_datetime(".txt"), "w") as f:
            f.write(text_data)
    
    def save_xlsx(self, datas):
        wb = Workbook()
        ws = wb.active
        border_style = Border(
            left = Side(border_style="thin", color='FF000000'),
            right = Side(border_style="thin", color='FF000000'),
            top = Side(border_style="thin", color='FF000000'),
            bottom = Side(border_style="thin", color='FF000000'),
        )
        r = 1
        for parent_row in datas:
            c = 1
            for child_row in parent_row:
                ws.cell(row=r, column=c).value = child_row
                ws.cell(row=r, column=c).border = border_style
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top")
                c += 1
            cellobj = None
            if self.hrefFlag == "yes":
                cellobj = ws.cell(row=r, column=6)
            else:
                cellobj = ws.cell(row=r, column=5)
            if cellobj.value == "適合":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FF40FFFF")
            elif cellobj.value == "適合(注記)":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FF40FF40")
            elif cellobj.value == "不適合":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FFFF8080")
            elif cellobj.value == "非適用":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FFC0C0C0")
            else:
                pass
            r += 1
        rangeobj = None
        if self.hrefFlag == "yes":
            rangeobj = range(1, 11)
        else:
            rangeobj = range(1, 10)
        for i in rangeobj:
            ws.cell(row=1, column=i).font = Font(bold=True)
            ws.cell(row=1, column=i).alignment = Alignment(horizontal="center")
        wb.save(self.fetch_filename_from_datetime(".xlsx"))


class LibraRepBotsUtil(LibraRepBots):
    def fetch_report_sequential(self, guideline_arr):
        rep_data = []
        self.wd.get(self.rep_index_url_base + self.projectID + "/")
        if guideline_arr is False:
            guideline_rows = self.open_text_data(self.guideline_filename)
        else:
            guideline_rows = guideline_arr
        page_rows = self.get_page_list_data()
        for guideline in guideline_rows:
            for pageID in page_rows:
                print(pageID, ",", guideline, "を処理しています。")
                path = self.fetch_report_detail_path(pageID, guideline)
                self.wd.get(path)
                time.sleep(self.shortWait)
                rep_data.extend(self.get_detail_table_data(pageID, guideline))
        add_header = None
        if self.hrefFlag == "yes":
            add_header = [["管理番号", "達成基準", "LibraURL", "状況/要件", "実装番号", "検査結果", "検査員", "コメント", "対象ソースコード", "修正ソースコード"]]
        else:
            add_header = [["管理番号", "達成基準", "状況/要件", "実装番号", "検査結果", "検査員", "コメント", "対象ソースコード", "修正ソースコード"]]
        last_rep_data = add_header + rep_data
        self.save_xlsx(last_rep_data)


args = sys.argv
projectID = args[1]
techID = None
try:
    techID = args[2]
except IndexError:
    pass
lbt = LibraRepBotsUtil(projectID)
time.sleep(lbt.shortWait)
lbt.login()
time.sleep(lbt.shortWait)
if techID is None:
    lbt.fetch_report_sequential(False)
else:
    lbt.fetch_report_sequential([techID])
time.sleep(lbt.shortWait)
lbt.logout
time.sleep(lbt.shortWait)
lbt.shutdown()